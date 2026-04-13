import io

from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import views

from apps.matches.models import Match
from apps.scoring.models import Ball
from apps.teams.models import Player, Team
from apps.tournament.models import Tournament


class ExportTournamentView(views.APIView):
    """Export full tournament data as a multi-sheet Excel file."""

    def get(self, request, tournament_id):
        tournament = Tournament.objects.get(
            id=tournament_id, created_by=request.user, is_deleted=False
        )

        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Tournament Name", tournament.name])
        ws.append(["Overs per Match", tournament.overs])
        ws.append(["Total Teams", tournament.total_teams])
        ws.append(["Players per Team", tournament.players_per_team])
        ws.append(["Pools", tournament.pool_count])
        ws.append(["Status", tournament.status])
        ws.append(["Created", str(tournament.created_at)])

        # --- Sheet 2: Teams ---
        ws_teams = wb.create_sheet("Teams")
        ws_teams.append(["Team", "Pool", "Player"])
        for team in Team.objects.filter(tournament=tournament).select_related("pool"):
            for player in team.players.all():
                ws_teams.append([team.name, team.pool.name if team.pool else "—", player.name])

        # --- Sheet 3: Match Results ---
        ws_matches = wb.create_sheet("Match Results")
        ws_matches.append([
            "Match #", "Stage", "Team 1", "Team 2", "Status",
            "Winner", "Result", "MOM",
        ])
        for m in Match.objects.filter(tournament=tournament).select_related(
            "team1", "team2", "winner", "mom_player"
        ):
            ws_matches.append([
                m.match_number, m.stage, m.team1.name, m.team2.name,
                m.status, m.winner.name if m.winner else "—",
                m.result_summary, m.mom_player.name if m.mom_player else "—",
            ])

        # --- Sheet 4: Batting Stats ---
        ws_bat = wb.create_sheet("Batting Stats")
        ws_bat.append(["Player", "Team", "Runs", "Balls Faced", "4s", "6s", "Strike Rate"])
        for player in Player.objects.filter(team__tournament=tournament):
            balls_faced = Ball.objects.filter(
                striker=player, is_wide=False
            )
            runs = sum(b.runs_scored for b in balls_faced)
            total_balls = balls_faced.count()
            fours = balls_faced.filter(runs_scored=4, is_boundary=True).count()
            sixes = balls_faced.filter(runs_scored=6, is_boundary=True).count()
            sr = round((runs / total_balls) * 100, 2) if total_balls > 0 else 0
            ws_bat.append([player.name, player.team.name, runs, total_balls, fours, sixes, sr])

        # --- Sheet 5: Bowling Stats ---
        ws_bowl = wb.create_sheet("Bowling Stats")
        ws_bowl.append(["Player", "Team", "Overs", "Runs Conceded", "Wickets", "Economy", "Wides", "No Balls"])
        for player in Player.objects.filter(team__tournament=tournament):
            bowled = Ball.objects.filter(bowler=player)
            if not bowled.exists():
                continue
            legal = bowled.filter(is_wide=False, is_noball=False).count()
            full_overs = legal // 6
            remaining = legal % 6
            overs_str = f"{full_overs}.{remaining}"
            runs_conceded = sum(b.total_runs for b in bowled)
            wickets = bowled.filter(is_wicket=True).exclude(wicket_type="RUN_OUT").count()
            overs_decimal = full_overs + (remaining / 6) if legal > 0 else 0
            economy = round(runs_conceded / overs_decimal, 2) if overs_decimal > 0 else 0
            wides = bowled.filter(is_wide=True).count()
            noballs = bowled.filter(is_noball=True).count()
            ws_bowl.append([
                player.name, player.team.name, overs_str,
                runs_conceded, wickets, economy, wides, noballs,
            ])

        # Write to response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{tournament.name.replace(' ', '_')}_export.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
